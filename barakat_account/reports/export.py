"""تصدير التقارير - Reports Export"""

import os
from datetime import datetime


def export_to_excel(transactions: list, filepath: str, company_name: str = "",
                    currency: str = "IQD") -> tuple:
    """تصدير الحركات إلى ملف Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return False, "يجب تثبيت مكتبة openpyxl: pip install openpyxl"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "كشف الحساب"

    # اتجاه RTL
    ws.sheet_view.rightToLeft = True

    # عنوان الشركة
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = company_name or "كشف الحساب"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    # تاريخ التصدير
    ws.merge_cells('A2:F2')
    date_cell = ws['A2']
    date_cell.value = f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    date_cell.alignment = Alignment(horizontal='center')
    date_cell.font = Font(size=11, color='666666')

    # رؤوس الجدول
    headers = ["التاريخ", "نوع الحركة", "المبلغ", "الرصيد بعد العملية", "الملاحظات", "رقم العملية"]
    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    header_font = Font(size=12, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # بيانات الحركات
    for row_idx, trans in enumerate(transactions, 5):
        values = [
            trans.get("date", ""),
            trans.get("type", ""),
            trans.get("amount", 0),
            trans.get("balance_after", 0),
            trans.get("notes", ""),
            trans.get("id", "")
        ]

        type_color = "E8F5E9" if trans.get("type") == "تعزيز" else "FFEBEE"
        row_fill = PatternFill(start_color=type_color, end_color=type_color, fill_type="solid")

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = row_fill
            if col_idx in (3, 4):
                cell.number_format = '#,##0'

    # عرض الأعمدة
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12

    try:
        wb.save(filepath)
        return True, f"تم التصدير بنجاح إلى:\n{filepath}"
    except Exception as e:
        return False, f"خطأ في التصدير: {str(e)}"


def export_to_pdf(transactions: list, filepath: str, company_name: str = "",
                  currency: str = "IQD") -> tuple:
    """تصدير الحركات إلى ملف PDF"""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm, mm
        from reportlab.lib.colors import HexColor, white, black
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return False, "يجب تثبيت مكتبة reportlab: pip install reportlab"

    try:
        pdfmetrics.registerFont(TTFont('Arabic', 'arial.ttf'))
        arabic_font = 'Arabic'
    except:
        arabic_font = 'Helvetica'

    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4),
                            rightMargin=2*cm, leftMargin=2*cm)

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Title'],
        fontName=arabic_font, fontSize=18, alignment=1,
        spaceAfter=10
    )

    elements.append(Paragraph(company_name or "كشف الحساب", title_style))
    elements.append(Paragraph(
        f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        ParagraphStyle('Date', parent=styles['Normal'], alignment=1,
                       fontSize=10, textColor=HexColor('#666666'))
    ))
    elements.append(Spacer(1, 20))

    table_data = [["التاريخ", "نوع الحركة", f"المبلغ ({currency})", "الرصيد بعد العملية", "الملاحظات"]]

    for trans in transactions:
        table_data.append([
            trans.get("date", ""),
            trans.get("type", ""),
            f"{trans.get('amount', 0):,.0f}",
            f"{trans.get('balance_after', 0):,.0f}",
            trans.get("notes", "")
        ])

    table = Table(table_data, colWidths=[4*cm, 3*cm, 4*cm, 5*cm, 6*cm])

    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#1a73e8')),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), arabic_font),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, HexColor('#f5f5f5')]),
    ])

    table.setStyle(table_style)
    elements.append(table)

    try:
        doc.build(elements)
        return True, f"تم التصدير بنجاح إلى:\n{filepath}"
    except Exception as e:
        return False, f"خطأ في التصدير: {str(e)}"
